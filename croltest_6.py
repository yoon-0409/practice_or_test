import requests
import openpyxl as xl
import time

# api 정보 셋팅
url = "https://dapi.kakao.com/v2/local/search/address.json"
REST_API_KEY = "bda47e44cef502cf8da02727209805b3"
headers = {"Authorization": "KakaoAK {}".format(REST_API_KEY)}


# 엑셀파일 읽기
wb = xl.load_workbook("샘플.xlsx", data_only=True)


list = ["아파트","빌라","더샵","타운","맨션","단지"]
def isBuilding(k,list):
  answer = False
  for i in list:
    if i in k:
      answer = True
  return answer

 
# 자료가 있는 시트 열기
ws = wb.worksheets[0]

# 루프돌면서 검증할 주소 읽기
for row, item in enumerate(ws.rows):
    if row > 0: #  두번째 행부터 
        addr = str(item[8].value)
        if "," in addr:
            addr = addr[:addr.index(",")]
        if "영일군" in addr:
            addr = addr.replace("영일군 ","")
        if "학전동" in addr:
            addr = addr.replace("동","리")
        if "자명동" in addr:
            addr = addr.replace("동","리")
        params = {"query": "{}".format(addr)} # 주소정보를 파라미터에  담습니다.
        resp = requests.get(url, params=params, headers=headers) # 해더와 파라미터 정보를 넣어 get 타입으로 전송합니다.

        documents = resp.json()["documents"] # json으로 받은 파일을 파싱하기.\
        # print(documents)
        if documents == []:
            ws.cell(row+1,14," ")
            ws.cell(row+1,14," ")
        else:
            address = documents[0]['address']['address_name']
            k = address.split(" ")
            if documents[0]['road_address']!=None:
                road_address = documents[0]['road_address']['address_name']
                k = documents[0]['road_address']['building_name']
                if isBuilding(k,list):
                    ws.cell(row+1,15,"집합건물")
                elif k == None: ws.cell(row+1,15," ")
                else:
                    ws.cell(row+1,15,"토지")
            else:
                ws.cell(row+1,15,"토지")
            #ws.cell(row+1, 14, road_address ) # 3번째 열에 값 엑셀쓰기
            ws.cell(row+1, 14, address)
 
        #time.sleep(0.3) # 너무 빨리 쿼리해서 죽을수도 있으니 쉬는 시간을 줍시다.
    if item[6].value == None: break

wb.save("샘플.xlsx") # 동일한 파일에 저장
wb.close() # 엑셀파일 다 썼으니 닫기.
